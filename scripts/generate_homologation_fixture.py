from __future__ import annotations

import argparse
import csv
import hashlib
import json
from datetime import UTC, datetime
from io import BytesIO, StringIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import Workbook

SEED = 5301
VERSION = "1.0.0"
FIXED_ZIP_TIME = (2026, 1, 1, 0, 0, 0)


def _stable_xlsx(workbook: Workbook) -> bytes:
    raw = BytesIO()
    workbook.save(raw)
    source = ZipFile(BytesIO(raw.getvalue()))
    target = BytesIO()
    with source, ZipFile(target, "w", ZIP_DEFLATED) as archive:
        for name in sorted(source.namelist()):
            original = source.getinfo(name)
            info = ZipInfo(name, FIXED_ZIP_TIME)
            info.compress_type = ZIP_DEFLATED
            info.external_attr = original.external_attr
            archive.writestr(info, source.read(name))
    return target.getvalue()


def workbook_bytes() -> bytes:
    workbook = Workbook()
    fixed_document_time = datetime(2026, 1, 1)
    workbook.properties.created = fixed_document_time
    workbook.properties.modified = fixed_document_time
    first = workbook.active
    first.title = "Inspecao_2026-09-01"
    first.append(["RELATORIO SINTETICO DE INSPECAO"])
    first.append([])
    first.append(
        [
            "lote_id",
            "produto",
            "linha",
            "turno",
            "status",
            "responsavel",
            "data",
            "observacao",
            "campo_adicional",
        ]
    )
    first.append(
        [
            "0000001",
            "MODEL-A",
            "L1",
            "A",
            "approved",
            "SYN-01",
            "01/09/2026",
            None,
            "preserved",
        ]
    )
    first.append(
        [
            "0000002",
            "MODEL-B",
            "L2",
            "B",
            "hold",
            "SYN-02",
            datetime(2026, 9, 1),
            "synthetic hold",
            None,
        ]
    )

    second = workbook.create_sheet("Inspecao_2026-09-02")
    second.append(["RELATORIO SINTETICO DE INSPECAO"])
    second.append([])
    second.append(["lote_id", "produto", "status", "data", "observacao"])
    second.append(["0000003", "MODEL-C", "pending", "02/09/2026", "synthetic"])

    reference = workbook.create_sheet("Base_Referencia")
    reference.append(["BASE SINTETICA DE REFERENCIA"])
    reference.append(
        ["lote_id", "codigo_produto", "descricao_produto", "status_cadastro"]
    )
    for index in range(1, 4):
        reference.append(
            [f"{index:07d}", f"MODEL-{chr(64 + index)}", "synthetic", "active"]
        )
    content = _stable_xlsx(workbook)
    workbook.close()
    return content


def plan_bytes() -> bytes:
    stream = StringIO(newline="")
    rows = [
        ["workorder_number", "lot_number", "model", "planned_quantity"],
        ["WO-SYN-001", "0000001", "MODEL-A", 1],
        ["WO-SYN-002", "0000002", "MODEL-B", 1],
        ["WO-SYN-003", "0000003", "MODEL-C", ""],
    ]
    csv.writer(stream, lineterminator="\n").writerows(rows)
    return stream.getvalue().encode("utf-8")


def generate(output: Path) -> dict:
    output.mkdir(parents=True, exist_ok=True)
    files = {
        "quality-reference.xlsx": workbook_bytes(),
        "plan-reference.csv": plan_bytes(),
    }
    entries = []
    for name, content in files.items():
        (output / name).write_bytes(content)
        entries.append(
            {
                "file": name,
                "bytes": len(content),
                "sha256": hashlib.sha256(content).hexdigest(),
            }
        )
    manifest = {
        "version": VERSION,
        "seed": SEED,
        "generated_at": datetime(2026, 9, 1, tzinfo=UTC).isoformat(),
        "contains_real_data": False,
        "sources": ["N-FP", "GMES/OQC"],
        "counts": {"workorders": 3, "quality_rows": 6, "sheets": 3},
        "files": entries,
    }
    (output / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    return manifest


def validate_manifest(path: Path) -> dict:
    manifest = json.loads(path.read_text(encoding="utf-8"))
    required = {
        "version", "seed", "generated_at", "contains_real_data",
        "sources", "counts", "files",
    }
    missing = required - manifest.keys()
    if missing:
        raise ValueError(f"Manifesto sem campos obrigatorios: {sorted(missing)}")
    if manifest.get("contains_real_data") is not False:
        raise ValueError("A fixture de homologacao deve declarar dados sinteticos")
    files = manifest["files"]
    if not isinstance(files, list) or not files:
        raise ValueError("Manifesto deve listar ao menos um arquivo")
    names = [item.get("file") for item in files if isinstance(item, dict)]
    if len(names) != len(files) or len(names) != len(set(names)):
        raise ValueError("Manifesto contem nomes ausentes ou duplicados")
    expected = {"file", "bytes", "sha256"}
    for item in files:
        if expected - item.keys():
            raise ValueError(f"Entrada de arquivo incompleta: {item!r}")
        relative = Path(item["file"])
        if relative.is_absolute() or relative.name != item["file"]:
            raise ValueError(f"Caminho de arquivo invalido: {item['file']}")
        file_path = path.parent / relative
        if not file_path.is_file():
            raise ValueError(f"Arquivo ausente: {item['file']}")
        content = file_path.read_bytes()
        if len(content) != item["bytes"]:
            raise ValueError(f"Tamanho divergente: {item['file']}")
        if hashlib.sha256(content).hexdigest() != item["sha256"]:
            raise ValueError(f"SHA-256 divergente: {item['file']}")
    present = {
        item.name for item in path.parent.iterdir() if item.is_file() and item != path
    }
    if present != set(names):
        raise ValueError("Artefatos da fixture divergem do manifesto")
    return manifest


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("data/synthetic/fixtures/homologated-workbook"),
    )
    args = parser.parse_args()
    manifest = generate(args.output)
    print(
        f"Fixture gerada: {manifest['counts']['workorders']} Workorders, "
        f"{manifest['counts']['sheets']} abas"
    )


if __name__ == "__main__":
    main()
