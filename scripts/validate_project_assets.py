from __future__ import annotations

import csv
import json
import os
import shutil
import subprocess
from pathlib import Path

from openpyxl import load_workbook

from generate_synthetic_data import validate_manifest

ROOT = Path(__file__).resolve().parents[1]
MIGRATIONS = ROOT / "database" / "migrations"
SYNTHETIC_DATA = ROOT / "data" / "synthetic"


def psql_command() -> list[str]:
    if shutil.which("psql"):
        return ["psql"]
    if not shutil.which("docker"):
        raise RuntimeError("psql e Docker Compose não estão disponíveis")
    return [
        "docker",
        "compose",
        "exec",
        "-T",
        "postgres",
        "psql",
        "--username",
        os.getenv("PGUSER", "synergia"),
        "--dbname",
        os.getenv("PGDATABASE", "synergia"),
    ]


def validate_migrations() -> None:
    migrations = sorted(MIGRATIONS.glob("*.sql"))
    if not migrations:
        raise ValueError("Nenhuma migration SQL foi encontrada")

    for migration in migrations:
        sql = migration.read_text(encoding="utf-8").strip()
        if not sql:
            raise ValueError(f"Migration vazia: {migration.relative_to(ROOT)}")
        subprocess.run(
            [
                *psql_command(),
                "--set=ON_ERROR_STOP=1",
                "--single-transaction",
            ],
            input=sql.encode("utf-8"),
            check=True,
        )
    print(f"Migrations SQL validadas: {len(migrations)}")


def validate_synthetic_data() -> None:
    supported = {".csv", ".json", ".xlsx"}
    files = [
        path
        for path in SYNTHETIC_DATA.rglob("*")
        if path.is_file() and path.suffix.lower() in supported
    ]

    for path in files:
        if path.name == "manifest.json":
            validate_manifest(path)
        elif path.suffix.lower() == ".json":
            json.loads(path.read_text(encoding="utf-8"))
        elif path.suffix.lower() == ".csv":
            with path.open(encoding="utf-8", newline="") as source:
                header = next(csv.reader(source), None)
                if not header or any(not column.strip() for column in header):
                    raise ValueError(f"CSV sem cabeçalho válido: {path.name}")
        else:
            workbook = load_workbook(path, read_only=True, data_only=False)
            try:
                header = next(workbook.active.iter_rows(values_only=True), None)
                if not header or any(
                    not str(column or "").strip() for column in header
                ):
                    raise ValueError(f"XLSX sem cabeçalho válido: {path.name}")
            finally:
                workbook.close()
    print(f"Arquivos sintéticos validados: {len(files)}")


if __name__ == "__main__":
    validate_migrations()
    validate_synthetic_data()
