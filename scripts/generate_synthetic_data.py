from __future__ import annotations

import argparse
import csv
import hashlib
import json
import random
import re
import zipfile
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import date, datetime, timedelta
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


GENERATOR_VERSION = "1.0.0"
SCHEMA_VERSION = "1.0"
SYNTHETIC_NOTICE = "SYNTHETIC DATA - NO REAL OR PERSONAL INFORMATION"
FIXED_ZIP_TIMESTAMP = (2026, 1, 1, 0, 0, 0)


@dataclass(frozen=True)
class Profile:
    workorders: int
    serials: int


PROFILES = {
    "minimal": Profile(workorders=4, serials=12),
    "small": Profile(workorders=50, serials=500),
    "medium": Profile(workorders=1_000, serials=12_000),
    "reference": Profile(workorders=6_800, serials=88_000),
}

SOURCES = ("N-FP", "OWM", "GMES/OQC", "TMS")
SOURCE_SLUGS = {
    "N-FP": "n-fp",
    "OWM": "owm",
    "GMES/OQC": "gmes-oqc",
    "TMS": "tms",
}
CANONICAL_FORMATS = {
    "N-FP": "xlsx",
    "OWM": "json",
    "GMES/OQC": "csv",
    "TMS": "json",
}
ALL_FORMATS = ("csv", "json", "xlsx")

HEADERS = {
    "N-FP": (
        "workorder_number",
        "demand_id",
        "lot_number",
        "model",
        "organization_code",
        "planned_quantity",
        "produced_quantity",
        "planned_date",
        "production_date",
        "status",
    ),
    "OWM": (
        "workorder_number",
        "demand_id",
        "lot_number",
        "serial_number",
        "model",
        "organization_code",
        "planned_quantity",
        "received_quantity",
        "released_quantity",
        "pending_quantity",
        "retained_quantity",
        "received_at",
        "released_at",
        "status",
    ),
    "GMES/OQC": (
        "workorder_number",
        "demand_id",
        "lot_number",
        "serial_number",
        "model",
        "organization_code",
        "decision_state",
        "oqc_flag",
        "hold_flag",
        "rework_flag",
        "hold_reason",
        "decided_at",
        "inspection_date",
    ),
    "TMS": (
        "workorder_number",
        "demand_id",
        "lot_number",
        "container_number",
        "organization_code",
        "quantity",
        "shipment_status",
        "ship_block_flag",
        "pending_reason",
        "shipment_date",
    ),
}


def _logical_digest(records: dict[str, list[dict[str, Any]]]) -> str:
    payload = json.dumps(
        records,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode()
    return hashlib.sha256(payload).hexdigest()


def _workorders(profile: Profile, rng: random.Random) -> list[dict[str, Any]]:
    base_date = date(2026, 1, 1)
    serial_counts = [profile.serials // profile.workorders] * profile.workorders
    for index in range(profile.serials % profile.workorders):
        serial_counts[index] += 1
    result: list[dict[str, Any]] = []
    for index, serial_count in enumerate(serial_counts, 1):
        planned = serial_count + rng.randint(1, 4)
        planned_date = base_date + timedelta(days=index % 180)
        result.append(
            {
                "workorder_number": f"SYN-WO-{index:06d}",
                "demand_id": f"SYN-DEM-{index:06d}",
                "lot_number": f"SYN-LOT-{index:06d}",
                "model": f"SYN-MODEL-{1 + (index - 1) % 20:02d}",
                "organization_code": f"SYN-ORG-{1 + (index - 1) % 8:03d}",
                "container_number": f"SYN-CONT-{1 + (index - 1) // 5:06d}",
                "planned_quantity": planned,
                "produced_quantity": serial_count,
                "serial_count": serial_count,
                "planned_date": planned_date.isoformat(),
                "production_date": (planned_date - timedelta(days=3)).isoformat(),
            }
        )
    return result


def build_dataset(
    *, profile_name: str, seed: int, scenario: str
) -> tuple[dict[str, list[dict[str, Any]]], dict[str, Any]]:
    if profile_name not in PROFILES:
        raise ValueError(f"Perfil desconhecido: {profile_name}")
    if scenario not in {"valid", "comprehensive"}:
        raise ValueError(f"Cenário desconhecido: {scenario}")
    profile = PROFILES[profile_name]
    rng = random.Random(seed)
    workorders = _workorders(profile, rng)
    n_fp: list[dict[str, Any]] = []
    owm: list[dict[str, Any]] = []
    gmes: list[dict[str, Any]] = []
    tms: list[dict[str, Any]] = []
    serial_index = 1
    for workorder in workorders:
        n_fp.append(
            {
                key: workorder[key]
                for key in (
                    "workorder_number",
                    "demand_id",
                    "lot_number",
                    "model",
                    "organization_code",
                    "planned_quantity",
                    "produced_quantity",
                    "planned_date",
                    "production_date",
                )
            }
            | {"status": "completed"}
        )
        for serial_offset in range(workorder["serial_count"]):
            serial = f"SYN-SER-{serial_index:08d}"
            released = 1
            received_at = date.fromisoformat(workorder["production_date"]) + timedelta(
                days=5 + serial_offset % 3
            )
            owm.append(
                {
                    "workorder_number": workorder["workorder_number"],
                    "demand_id": workorder["demand_id"],
                    "lot_number": workorder["lot_number"],
                    "serial_number": serial,
                    "model": workorder["model"],
                    "organization_code": workorder["organization_code"],
                    "planned_quantity": "",
                    "received_quantity": 1,
                    "released_quantity": released,
                    "pending_quantity": 0,
                    "retained_quantity": 0,
                    "received_at": received_at.isoformat(),
                    "released_at": (received_at + timedelta(days=1)).isoformat(),
                    "status": "released",
                }
            )
            gmes.append(
                {
                    "workorder_number": workorder["workorder_number"],
                    "demand_id": workorder["demand_id"],
                    "lot_number": workorder["lot_number"],
                    "serial_number": serial,
                    "model": workorder["model"],
                    "organization_code": workorder["organization_code"],
                    "decision_state": "oqc_pass",
                    "oqc_flag": True,
                    "hold_flag": False,
                    "rework_flag": False,
                    "hold_reason": "",
                    "decided_at": (received_at + timedelta(days=1)).isoformat(),
                    "inspection_date": received_at.isoformat(),
                }
            )
            serial_index += 1
        tms.append(
            {
                "workorder_number": workorder["workorder_number"],
                "demand_id": workorder["demand_id"],
                "lot_number": workorder["lot_number"],
                "container_number": workorder["container_number"],
                "organization_code": workorder["organization_code"],
                "quantity": workorder["serial_count"],
                "shipment_status": "released",
                "ship_block_flag": False,
                "pending_reason": "",
                "shipment_date": workorder["planned_date"],
            }
        )

    scenario_counts: Counter[str] = Counter(
        {"fully_valid": sum(map(len, (n_fp, owm, gmes, tms)))}
    )
    expected_validation_codes: Counter[str] = Counter()
    expected_processing_codes: Counter[str] = Counter()
    expected_rule_ids: set[str] = {"oqc_pass"}
    if scenario == "comprehensive":
        owm[0]["planned_quantity"] = n_fp[0]["planned_quantity"] + 7
        scenario_counts["source_divergence"] += 1
        expected_processing_codes["source_divergence"] += 1
        expected_rule_ids.add("source_divergence")

        absent = dict(owm[1])
        absent.update(
            workorder_number="SYN-WO-ABSENT",
            demand_id="",
            lot_number="",
            serial_number="",
            planned_quantity="",
            received_quantity=1,
            released_quantity=0,
            pending_quantity=1,
            retained_quantity=0,
            status="pending",
        )
        owm.append(absent)
        scenario_counts["workorder_absent"] += 1
        expected_processing_codes["missing_source_match"] += 1

        gmes[2]["lot_number"] = ""
        scenario_counts["lot_absent"] += 1

        owm[2]["serial_number"] = "SYN-SERIAL@INVALID"
        scenario_counts["invalid_serial"] += 1
        expected_validation_codes["invalid_identifier"] += 1

        owm[4]["serial_number"] = owm[3]["serial_number"]
        scenario_counts["duplicate_serial"] += 1
        expected_validation_codes["duplicate_serial"] += 1

        n_fp[1]["organization_code"] = "SYN-ORG-UNKNOWN"
        scenario_counts["unknown_organization"] += 1
        expected_validation_codes["unknown_organization"] += 1

        first_workorder = workorders[0]["workorder_number"]
        for row in owm:
            if row["workorder_number"] == first_workorder:
                row["released_quantity"] = 0
                row["pending_quantity"] = 1
                row["status"] = "pending"
                break
        scenario_counts["partial_release"] += 1

        gmes[0].update(decision_state="oqc_pending", oqc_flag=False)
        scenario_counts["oqc_pending"] += 1
        expected_rule_ids.add("oqc_pending")

        gmes[1].update(
            decision_state="oqc_hold",
            oqc_flag=False,
            hold_flag=True,
            hold_reason="SYNTHETIC OQC HOLD",
        )
        scenario_counts["oqc_hold"] += 1
        expected_rule_ids.add("oqc_hold")

        gmes[5].update(
            decision_state="long_term_hold",
            oqc_flag=False,
            hold_flag=True,
            hold_reason="SYNTHETIC LONG TERM HOLD",
            decided_at="2025-01-01",
        )
        scenario_counts["long_term_hold"] += 1
        expected_rule_ids.add("long_term_hold")

        gmes[6].update(
            decision_state="rework",
            oqc_flag=False,
            rework_flag=True,
            hold_reason="SYNTHETIC REWORK",
        )
        scenario_counts["rework"] += 1
        expected_rule_ids.add("rework")

        tms[0].update(
            shipment_status="ship_block",
            ship_block_flag=True,
            pending_reason="SYNTHETIC SHIP BLOCK",
        )
        scenario_counts["ship_block"] += 1
        expected_rule_ids.add("ship_block")

        n_fp[2]["workorder_number"] = ""
        scenario_counts["required_data_absent"] += 1
        expected_validation_codes["required_field"] += 1

        n_fp[3]["planned_quantity"] = -1
        n_fp[3]["planned_date"] = "2026-99-99"
        scenario_counts["invalid_date_and_quantity"] += 1
        expected_validation_codes["invalid_quantity"] += 1
        expected_validation_codes["invalid_date"] += 1

    records = {"N-FP": n_fp, "OWM": owm, "GMES/OQC": gmes, "TMS": tms}
    metadata = {
        "profile": asdict(profile),
        "entities": {
            "workorders": profile.workorders,
            "lots": profile.workorders,
            "serials": profile.serials,
            "organizations": min(profile.workorders, 8),
        },
        "scenario_counts": dict(sorted(scenario_counts.items())),
        "expected_validation_codes": dict(sorted(expected_validation_codes.items())),
        "expected_processing_codes": dict(sorted(expected_processing_codes.items())),
        "expected_rule_ids": sorted(expected_rule_ids),
        "known_organizations": [f"SYN-ORG-{index:03d}" for index in range(1, 9)],
    }
    return records, metadata


def _csv_bytes(headers: tuple[str, ...], records: list[dict[str, Any]]) -> bytes:
    stream = StringIO(newline="")
    writer = csv.DictWriter(
        stream,
        fieldnames=headers,
        extrasaction="ignore",
        lineterminator="\n",
    )
    writer.writeheader()
    for record in records:
        writer.writerow(
            {
                header: (
                    "true"
                    if record.get(header) is True
                    else "false"
                    if record.get(header) is False
                    else record.get(header, "")
                )
                for header in headers
            }
        )
    return stream.getvalue().encode()


def _json_bytes(records: list[dict[str, Any]]) -> bytes:
    return (json.dumps(records, ensure_ascii=False, indent=2) + "\n").encode()


def _deterministic_xlsx_bytes(
    headers: tuple[str, ...], records: list[dict[str, Any]]
) -> bytes:
    workbook = Workbook(write_only=True)
    workbook.properties.creator = "SYNERGIA synthetic generator"
    workbook.properties.title = SYNTHETIC_NOTICE
    workbook.properties.created = datetime(2026, 1, 1)
    workbook.properties.modified = datetime(2026, 1, 1)
    worksheet = workbook.create_sheet("Synthetic")
    worksheet.append(list(headers))
    for record in records:
        worksheet.append([record.get(header, "") for header in headers])
    raw = BytesIO()
    workbook.save(raw)
    workbook.close()

    normalized = BytesIO()
    with (
        zipfile.ZipFile(raw) as source,
        zipfile.ZipFile(
            normalized, "w", zipfile.ZIP_DEFLATED, compresslevel=9
        ) as target,
    ):
        for name in sorted(source.namelist()):
            info = zipfile.ZipInfo(name, FIXED_ZIP_TIMESTAMP)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = 0o600 << 16
            content = source.read(name)
            if name == "docProps/core.xml":
                content = re.sub(
                    rb"<dcterms:modified[^>]*>.*?</dcterms:modified>",
                    (
                        b'<dcterms:modified xsi:type="dcterms:W3CDTF">'
                        b"2026-01-01T00:00:00Z</dcterms:modified>"
                    ),
                    content,
                )
            target.writestr(info, content, compress_type=zipfile.ZIP_DEFLATED)
    return normalized.getvalue()


def _formats_for(source: str, formats: str) -> tuple[str, ...]:
    if formats == "canonical":
        return (CANONICAL_FORMATS[source],)
    if formats == "all":
        return ALL_FORMATS
    selected = tuple(
        dict.fromkeys(value.strip().lower() for value in formats.split(","))
    )
    if not selected or any(value not in ALL_FORMATS for value in selected):
        raise ValueError("Formatos aceitos: canonical, all, csv, json e xlsx")
    return selected


def _write_file(
    path: Path, source: str, extension: str, records: list[dict[str, Any]]
) -> None:
    headers = HEADERS[source]
    content = (
        _csv_bytes(headers, records)
        if extension == "csv"
        else _json_bytes(records)
        if extension == "json"
        else _deterministic_xlsx_bytes(headers, records)
    )
    path.write_bytes(content)


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        while chunk := stream.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def _read_count(path: Path) -> int:
    if path.suffix == ".json":
        document = json.loads(path.read_text(encoding="utf-8"))
        return len(document if isinstance(document, list) else [document])
    if path.suffix == ".csv":
        with path.open(encoding="utf-8", newline="") as stream:
            return max(sum(1 for _ in csv.reader(stream)) - 1, 0)
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        return max(sum(1 for _ in workbook.active.iter_rows()) - 1, 0)
    finally:
        workbook.close()


def generate_bundle(
    *,
    output: Path,
    profile_name: str,
    seed: int,
    scenario: str,
    formats: str = "canonical",
) -> dict[str, Any]:
    if output.exists() and any(output.iterdir()):
        raise FileExistsError(f"Diretório de saída não está vazio: {output}")
    output.mkdir(parents=True, exist_ok=True)
    records, metadata = build_dataset(
        profile_name=profile_name, seed=seed, scenario=scenario
    )
    files: list[dict[str, Any]] = []
    for source in SOURCES:
        for extension in _formats_for(source, formats):
            name = f"{SOURCE_SLUGS[source]}.{extension}"
            path = output / name
            _write_file(path, source, extension, records[source])
            files.append(
                {
                    "source": source,
                    "format": extension,
                    "file": name,
                    "records": len(records[source]),
                    "size_bytes": path.stat().st_size,
                    "sha256": _file_sha256(path),
                }
            )
    manifest = {
        "schema_version": SCHEMA_VERSION,
        "generator_version": GENERATOR_VERSION,
        "synthetic_notice": SYNTHETIC_NOTICE,
        "seed": seed,
        "profile": profile_name,
        "scenario": scenario,
        "logical_digest": _logical_digest(records),
        "configuration": metadata["profile"],
        "entities": metadata["entities"],
        "sources": {source: {"records": len(records[source])} for source in SOURCES},
        "files": files,
        "scenario_counts": metadata["scenario_counts"],
        "expectations": {
            "validation_issue_codes": metadata["expected_validation_codes"],
            "processing_issue_codes": metadata["expected_processing_codes"],
            "classification_rule_ids": metadata["expected_rule_ids"],
            "known_organizations": metadata["known_organizations"],
            "valid_pipeline": (
                {
                    "rows_read": sum(len(records[source]) for source in SOURCES),
                    "valid_records": sum(len(records[source]) for source in SOURCES),
                    "rejected_records": 0,
                    "normalized_records": sum(
                        len(records[source]) for source in SOURCES
                    ),
                    "consolidated_workorders": metadata["entities"]["workorders"],
                    "consolidated_lots": metadata["entities"]["lots"],
                    "consolidated_serials": metadata["entities"]["serials"],
                }
                if scenario == "valid"
                else None
            ),
        },
    }
    manifest_path = output / "manifest.json"
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    validate_manifest(manifest_path)
    return manifest


def validate_manifest(path: Path) -> dict[str, Any]:
    manifest = json.loads(path.read_text(encoding="utf-8"))
    required = {
        "schema_version",
        "generator_version",
        "synthetic_notice",
        "seed",
        "profile",
        "scenario",
        "logical_digest",
        "configuration",
        "entities",
        "sources",
        "files",
        "scenario_counts",
        "expectations",
    }
    missing = required - manifest.keys()
    if missing:
        raise ValueError(f"Manifesto sem campos obrigatórios: {sorted(missing)}")
    if manifest["synthetic_notice"] != SYNTHETIC_NOTICE:
        raise ValueError("Manifesto sem identificação sintética oficial")
    for item in manifest["files"]:
        file_path = path.parent / item["file"]
        if not file_path.is_file():
            raise ValueError(f"Arquivo do manifesto ausente: {item['file']}")
        digest = _file_sha256(file_path)
        if digest != item["sha256"]:
            raise ValueError(f"SHA-256 divergente: {item['file']}")
        if _read_count(file_path) != item["records"]:
            raise ValueError(f"Contagem divergente: {item['file']}")
    return manifest


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Gera massas sintéticas determinísticas do SYNERGIA"
    )
    parser.add_argument("--profile", choices=PROFILES, default="small")
    parser.add_argument(
        "--scenario", choices=("valid", "comprehensive"), default="valid"
    )
    parser.add_argument("--seed", type=int, default=20260830)
    parser.add_argument("--formats", default="canonical")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("artifacts/synthetic"),
        help="Diretório novo ou vazio para os arquivos gerados",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    manifest = generate_bundle(
        output=args.output,
        profile_name=args.profile,
        seed=args.seed,
        scenario=args.scenario,
        formats=args.formats,
    )
    print(
        f"Gerado perfil={manifest['profile']} cenário={manifest['scenario']} "
        f"seed={manifest['seed']} arquivos={len(manifest['files'])} "
        f"digest={manifest['logical_digest']}"
    )


if __name__ == "__main__":
    main()
