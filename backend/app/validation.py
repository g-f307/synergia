from __future__ import annotations

import csv
import json
import re
import unicodedata
from collections.abc import Collection
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class SourceSchema:
    required: tuple[str, ...]
    aliases: dict[str, str]
    quantities: tuple[str, ...] = ()
    dates: tuple[str, ...] = ()


COMMON_ALIASES = {
    "workorder": "workorder_number",
    "work_order": "workorder_number",
    "wo": "workorder_number",
    "serial": "serial_number",
    "serial_no": "serial_number",
    "organization": "organization_code",
    "organisation": "organization_code",
    "org": "organization_code",
    "lote": "lot_number",
    "lote_id": "lot_number",
    "lot": "lot_number",
    "produto": "model",
    "codigo_produto": "model",
    "data": "inspection_date",
    "observacao": "reason",
}

SCHEMAS = {
    "N-FP": SourceSchema(
        ("workorder_number",),
        COMMON_ALIASES,
        ("planned_quantity", "produced_quantity"),
        ("planned_date", "production_date"),
    ),
    "OWM": SourceSchema(
        ("workorder_number",),
        COMMON_ALIASES,
        (
            "received_quantity",
            "released_quantity",
            "pending_quantity",
            "retained_quantity",
        ),
        ("received_at", "released_at"),
    ),
    "GMES/OQC": SourceSchema((), COMMON_ALIASES, (), ("decided_at", "inspection_date")),
    "TMS": SourceSchema(
        ("workorder_number",), COMMON_ALIASES, ("quantity",), ("shipment_date",)
    ),
}

FORMULA_ERRORS = ("#VALUE!", "#REF!", "#DIV/0!", "#NAME?", "#NUM!", "#N/A", "#NULL!")
IDENTIFIER_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._/-]*$")


class DataReadError(Exception):
    def __init__(self, reason: str, *, sheet: str, row: int | None = None) -> None:
        super().__init__(reason)
        self.reason = reason
        self.sheet = sheet
        self.row = row


class LocatedRow(list[Any]):
    """Spreadsheet row that retains its physical position after preambles."""

    def __init__(self, values: list[Any], row_number: int) -> None:
        super().__init__(values)
        self.row_number = row_number


RELATIONSHIP_IDENTIFIERS = (
    "workorder_number",
    "demand_id",
    "serial_number",
    "lot_number",
)


def _header(value: Any) -> str:
    text = "".join(
        character
        for character in unicodedata.normalize(
            "NFKD", str(value or "").strip().lower()
        )
        if not unicodedata.combining(character)
    )
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def _header_candidates(source: str | None) -> set[str]:
    schemas = [SCHEMAS[source]] if source else list(SCHEMAS.values())
    candidates = set(RELATIONSHIP_IDENTIFIERS)
    for schema in schemas:
        candidates.update(schema.required)
        candidates.update(schema.aliases)
        candidates.update(schema.quantities)
        candidates.update(schema.dates)
    candidates.update({"status", "reason", "model", "oqc_flag"})
    return candidates


def _xlsx_header_index(rows: list[list[Any]], source: str | None) -> int | None:
    """Find the first credible table header and ignore presentation preambles."""
    candidates = _header_candidates(source)
    schema = SCHEMAS.get(source) if source else None
    for index, row in enumerate(rows):
        normalized = [_header(value) for value in row if value not in (None, "")]
        if len(normalized) < 2 or len(normalized) != len(set(normalized)):
            continue
        canonical = (
            [schema.aliases.get(value, value) for value in normalized]
            if schema
            else normalized
        )
        recognized = {value for value in normalized if value in candidates}
        has_relationship = any(
            value in RELATIONSHIP_IDENTIFIERS for value in canonical
        )
        if has_relationship and len(recognized) >= 2:
            return index
    return None


def _identifier_key(value: Any) -> str:
    return re.sub(r"\s+", "", str(value).strip()).upper()


def _issue(
    code: str,
    reason: str,
    *,
    severity: str = "error",
    sheet: str | None = None,
    row: int | None = None,
    column: str | None = None,
) -> dict[str, Any]:
    return {
        "severity": severity,
        "code": code,
        "sheet": sheet,
        "row": row,
        "column": column,
        "reason": reason,
    }


def _read(
    path: Path, extension: str, source: str | None = None
) -> tuple[list[tuple[str, list[str], list[list[Any]]]], list[dict[str, Any]]]:
    issues: list[dict[str, Any]] = []
    tables: list[tuple[str, list[str], list[list[Any]]]] = []
    if extension == ".csv":
        with path.open(encoding="utf-8-sig", newline="") as stream:
            reader = csv.reader(stream, strict=True)
            try:
                values = list(reader)
            except csv.Error as exc:
                raise DataReadError(
                    "CSV malformado durante a leitura",
                    sheet="CSV",
                    row=reader.line_num or None,
                ) from exc
        if values:
            tables.append(("CSV", values[0], values[1:]))
    elif extension == ".json":
        with path.open(encoding="utf-8") as stream:
            document = json.load(stream)
        records = document if isinstance(document, list) else [document]
        if records and all(isinstance(item, dict) for item in records):
            headers = list(dict.fromkeys(key for item in records for key in item))
            tables.append(
                (
                    "JSON",
                    headers,
                    [[item.get(key) for key in headers] for item in records],
                )
            )
        else:
            issues.append(
                _issue(
                    "invalid_structure",
                    "JSON deve conter um objeto ou uma lista de objetos",
                    sheet="JSON",
                )
            )
    else:
        with path.open("rb") as stream:
            workbook = load_workbook(stream, read_only=True, data_only=False)
            try:
                for worksheet in workbook.worksheets:
                    values = [
                        list(row) for row in worksheet.iter_rows(values_only=True)
                    ]
                    header_index = _xlsx_header_index(values, source)
                    if header_index is None:
                        issues.append(
                            _issue(
                                "invalid_header",
                                "Aba vazia ou sem cabeçalho",
                                sheet=worksheet.title,
                            )
                        )
                        continue
                    tables.append(
                        (
                            worksheet.title,
                            values[header_index],
                            [
                                LocatedRow(row, row_number)
                                for row_number, row in enumerate(
                                    values[header_index + 1 :], header_index + 2
                                )
                            ],
                        )
                    )
            finally:
                workbook.close()
    return tables, issues


read_tables = _read


def _valid_date(value: Any) -> bool:
    if isinstance(value, date | datetime):
        return True
    if not isinstance(value, str) or not value.strip():
        return False
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%dT%H:%M:%S"):
        try:
            datetime.strptime(value.strip(), fmt)
            return True
        except ValueError:
            pass
    return False


def _report(
    source: str, row_count: int, issues: list[dict[str, Any]]
) -> dict[str, Any]:
    errors = sum(item["severity"] == "error" for item in issues)
    warnings = sum(item["severity"] == "warning" for item in issues)
    return {
        "source": source,
        "valid": errors == 0,
        "blocking": errors > 0,
        "row_count": row_count,
        "error_count": errors,
        "warning_count": warnings,
        "issues": issues,
    }


def failed_validation_report(
    source: str,
    code: str,
    reason: str,
    *,
    sheet: str | None = None,
    row: int | None = None,
) -> dict[str, Any]:
    return _report(source, 0, [_issue(code, reason, sheet=sheet, row=row)])


def validate_tables(
    tables: list[tuple[str, list[str], list[list[Any]]]],
    source: str,
    known_organizations: Collection[str] | None = None,
    initial_issues: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    schema = SCHEMAS[source]
    allowed_organizations = (
        {value.strip().upper() for value in known_organizations}
        if known_organizations is not None
        else None
    )
    issues = list(initial_issues or [])
    seen_serials: dict[str, tuple[str, int]] = {}
    row_count = 0
    if not tables:
        issues.append(_issue("empty_file", "Arquivo não contém linhas de dados"))

    for sheet, raw_headers, rows in tables:
        seen_rows: dict[tuple[str, ...], int] = {}
        normalized = [_header(value) for value in raw_headers]
        headers = [schema.aliases.get(value, value) for value in normalized]
        header_row = (
            max(1, getattr(rows[0], "row_number", 2) - 1) if rows else 1
        )
        for index, value in enumerate(normalized, 1):
            if not value:
                issues.append(
                    _issue(
                        "invalid_header",
                        "Cabeçalho vazio",
                        sheet=sheet,
                        row=header_row,
                        column=get_column_letter(index),
                    )
                )
        for required in schema.required:
            if required not in headers:
                issues.append(
                    _issue(
                        "missing_column",
                        f"Coluna obrigatória ausente: {required}",
                        sheet=sheet,
                        row=header_row,
                        column=required,
                    )
                )
        if source == "GMES/OQC" and not any(
            field in headers for field in RELATIONSHIP_IDENTIFIERS
        ):
            issues.append(
                _issue(
                    "missing_column",
                    "Informe Workorder, Demand ID, serial ou lote",
                    sheet=sheet,
                    row=header_row,
                )
            )
        for offset, values in enumerate(rows, 2):
            offset = getattr(values, "row_number", offset)
            if not any(value not in (None, "") for value in values):
                issues.append(
                    _issue(
                        "empty_row",
                        "Linha vazia",
                        severity="warning",
                        sheet=sheet,
                        row=offset,
                    )
                )
                continue
            row_count += 1
            row_key = tuple(str(value).strip() for value in values)
            if row_key in seen_rows:
                issues.append(
                    _issue(
                        "duplicate_row",
                        "Linha duplicada; primeira ocorrência na linha "
                        f"{seen_rows[row_key]}",
                        severity="warning",
                        sheet=sheet,
                        row=offset,
                    )
                )
            else:
                seen_rows[row_key] = offset
            record = {
                header: values[index] if index < len(values) else None
                for index, header in enumerate(headers)
            }
            for index, value in enumerate(values, 1):
                text = str(value).upper() if value is not None else ""
                formula_error = next(
                    (error for error in FORMULA_ERRORS if error in text), None
                )
                if formula_error:
                    code = (
                        "broken_reference"
                        if formula_error == "#REF!"
                        else "invalid_formula"
                    )
                    issues.append(
                        _issue(
                            code,
                            f"Erro de fórmula preservado: {formula_error}",
                            sheet=sheet,
                            row=offset,
                            column=get_column_letter(index),
                        )
                    )
            for field in schema.required:
                if field in headers and record.get(field) in (None, ""):
                    issues.append(
                        _issue(
                            "required_field",
                            f"Campo obrigatório vazio: {field}",
                            sheet=sheet,
                            row=offset,
                            column=get_column_letter(headers.index(field) + 1),
                        )
                    )
            if source == "GMES/OQC" and not any(
                record.get(field) not in (None, "")
                for field in RELATIONSHIP_IDENTIFIERS
                if field in headers
            ):
                issues.append(
                    _issue(
                        "required_field",
                        "Informe ao menos um identificador de relacionamento",
                        sheet=sheet,
                        row=offset,
                    )
                )
            for field in schema.quantities:
                value = record.get(field)
                if value in (None, ""):
                    continue
                valid = not isinstance(value, bool)
                try:
                    number = float(value)
                    valid = valid and number >= 0 and number.is_integer()
                except (TypeError, ValueError):
                    valid = False
                if not valid:
                    issues.append(
                        _issue(
                            "invalid_quantity",
                            f"Quantidade inválida: {field}",
                            sheet=sheet,
                            row=offset,
                            column=get_column_letter(headers.index(field) + 1),
                        )
                    )
            for field in schema.dates:
                value = record.get(field)
                if value not in (None, "") and not _valid_date(value):
                    issues.append(
                        _issue(
                            "invalid_date",
                            f"Data inválida: {field}",
                            sheet=sheet,
                            row=offset,
                            column=get_column_letter(headers.index(field) + 1),
                        )
                    )
            for field in ("workorder_number", "serial_number"):
                value = record.get(field)
                if value not in (None, "") and not IDENTIFIER_RE.fullmatch(
                    _identifier_key(value)
                ):
                    issues.append(
                        _issue(
                            "invalid_identifier",
                            f"Identificador inválido: {field}",
                            sheet=sheet,
                            row=offset,
                            column=get_column_letter(headers.index(field) + 1),
                        )
                    )
            serial = record.get("serial_number")
            if serial not in (None, ""):
                serial_key = _identifier_key(serial)
                if serial_key in seen_serials:
                    first_sheet, first_row = seen_serials[serial_key]
                    issues.append(
                        _issue(
                            "duplicate_serial",
                            "Serial duplicado; primeira ocorrência em "
                            f"{first_sheet}:{first_row}",
                            sheet=sheet,
                            row=offset,
                            column=get_column_letter(
                                headers.index("serial_number") + 1
                            ),
                        )
                    )
                else:
                    seen_serials[serial_key] = (sheet, offset)
            organization = record.get("organization_code")
            if (
                allowed_organizations is not None
                and organization not in (None, "")
                and str(organization).strip().upper() not in allowed_organizations
            ):
                issues.append(
                    _issue(
                        "unknown_organization",
                        "Organização desconhecida",
                        sheet=sheet,
                        row=offset,
                        column=get_column_letter(
                            headers.index("organization_code") + 1
                        ),
                    )
                )
            parent = record.get("workorder_reference")
            if (
                parent not in (None, "")
                and str(parent).strip()
                != str(record.get("workorder_number") or "").strip()
            ):
                issues.append(
                    _issue(
                        "unmatched_key",
                        "Chave sem correspondência no Workorder da linha",
                        sheet=sheet,
                        row=offset,
                        column=get_column_letter(
                            headers.index("workorder_reference") + 1
                        ),
                    )
                )

    if row_count == 0 and not any(item["code"] == "empty_file" for item in issues):
        issues.append(_issue("empty_file", "Arquivo não contém linhas de dados"))
    return _report(source, row_count, issues)


def validate_file(
    path: Path,
    extension: str,
    source: str,
    known_organizations: Collection[str] | None = None,
) -> dict[str, Any]:
    try:
        tables, issues = read_tables(path, extension, source)
    except DataReadError as exc:
        return failed_validation_report(
            source,
            "read_error",
            exc.reason,
            sheet=exc.sheet,
            row=exc.row,
        )
    return validate_tables(tables, source, known_organizations, issues)
