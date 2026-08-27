from __future__ import annotations

import csv
import hashlib
import json
import logging
import os
import shutil
from collections.abc import Generator
from datetime import datetime
from enum import Enum
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Annotated, Protocol
from uuid import uuid4

import psycopg
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile, status
from openpyxl import load_workbook
from psycopg.rows import dict_row
from pydantic import BaseModel

from app.validation import failed_validation_report, validate_file

logger = logging.getLogger("synergia.imports")
router = APIRouter(prefix="/imports", tags=["imports"])


class ImportSource(str, Enum):
    n_fp = "N-FP"
    owm = "OWM"
    gmes_oqc = "GMES/OQC"
    tms = "TMS"


class ImportStatus(BaseModel):
    execution_id: str
    status: str
    source: ImportSource
    file_name: str | None = None
    extension: str | None = None
    size_bytes: int | None = None
    sha256: str | None = None
    actor_type: str
    actor_identifier: str
    started_at: datetime
    finished_at: datetime | None
    failure_reason: str | None = None
    duplicate_of_execution_id: str | None = None


class ValidationReport(BaseModel):
    execution_id: str
    source: ImportSource
    file_name: str
    valid: bool
    blocking: bool
    row_count: int
    error_count: int
    warning_count: int
    issues: list[dict]


class ImportRepository(Protocol):
    def start(
        self, execution_id: str, source: str, actor_type: str, actor: str
    ) -> None: ...

    def claim_file(
        self,
        execution_id: str,
        *,
        file_name: str,
        extension: str,
        size_bytes: int,
        digest: str,
        media_type: str | None,
        storage_key: str,
    ) -> str | None: ...

    def mark_completed(self, execution_id: str) -> None: ...

    def mark_validation_failed(self, execution_id: str) -> None: ...

    def abort_claim(self, execution_id: str, reason: str) -> None: ...

    def finish(
        self,
        execution_id: str,
        state: str,
        reason: str,
        duplicate_of: str | None = None,
    ) -> None: ...

    def get(self, execution_id: str) -> dict | None: ...


class PostgresImportRepository:
    def __init__(self, database_url: str) -> None:
        self.database_url = database_url

    def _connect(self):
        return psycopg.connect(self.database_url, row_factory=dict_row)

    def start(
        self, execution_id: str, source: str, actor_type: str, actor: str
    ) -> None:
        with self._connect() as connection:
            connection.execute(
                """
                INSERT INTO synergia.executions
                    (id, status, source, actor_type, actor_identifier)
                VALUES (%s, 'running', %s, %s, %s)
                """,
                (execution_id, source, actor_type, actor),
            )

    def claim_file(
        self,
        execution_id: str,
        *,
        file_name: str,
        extension: str,
        size_bytes: int,
        digest: str,
        media_type: str | None,
        storage_key: str,
    ) -> str | None:
        with self._connect() as connection:
            inserted = connection.execute(
                """
                INSERT INTO synergia.source_files
                    (execution_id, file_name, extension, content_hash, media_type,
                     size_bytes, storage_key)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (content_hash) DO NOTHING
                RETURNING execution_id
                """,
                (
                    execution_id,
                    file_name,
                    extension,
                    digest,
                    media_type,
                    size_bytes,
                    storage_key,
                ),
            ).fetchone()
            if inserted:
                return None

            duplicate = connection.execute(
                """
                SELECT execution_id
                FROM synergia.source_files
                WHERE content_hash = %s
                """,
                (digest,),
            ).fetchone()
            if duplicate is None:
                raise RuntimeError("Conflito de hash sem execução original")
            return duplicate["execution_id"]

    def mark_completed(self, execution_id: str) -> None:
        with self._connect() as connection:
            connection.execute(
                """
                UPDATE synergia.executions
                SET status = 'completed', finished_at = now(), updated_at = now()
                WHERE id = %s
                """,
                (execution_id,),
            )

    def mark_validation_failed(self, execution_id: str) -> None:
        with self._connect() as connection:
            connection.execute(
                """
                UPDATE synergia.executions
                SET status = 'validation_failed', failure_reason = 'validation_failed',
                    finished_at = now(), updated_at = now()
                WHERE id = %s
                """,
                (execution_id,),
            )

    def abort_claim(self, execution_id: str, reason: str) -> None:
        with self._connect() as connection:
            connection.execute(
                "DELETE FROM synergia.source_files WHERE execution_id = %s",
                (execution_id,),
            )
            connection.execute(
                """
                UPDATE synergia.executions
                SET status = 'failed', failure_reason = %s,
                    finished_at = now(), updated_at = now()
                WHERE id = %s
                """,
                (reason, execution_id),
            )

    def finish(
        self,
        execution_id: str,
        state: str,
        reason: str,
        duplicate_of: str | None = None,
    ) -> None:
        with self._connect() as connection:
            connection.execute(
                """
                UPDATE synergia.executions
                SET status = %s, failure_reason = %s,
                    duplicate_of_execution_id = %s,
                    finished_at = now(), updated_at = now()
                WHERE id = %s
                """,
                (state, reason, duplicate_of, execution_id),
            )

    def get(self, execution_id: str) -> dict | None:
        with self._connect() as connection:
            return connection.execute(
                """
                SELECT e.id AS execution_id, e.status, e.source,
                       sf.file_name, sf.extension, sf.size_bytes,
                       sf.content_hash AS sha256, e.actor_type,
                       e.actor_identifier, e.started_at, e.finished_at,
                       e.failure_reason, e.duplicate_of_execution_id
                FROM synergia.executions e
                LEFT JOIN synergia.source_files sf ON sf.execution_id = e.id
                WHERE e.id = %s
                """,
                (execution_id,),
            ).fetchone()


def get_repository() -> Generator[ImportRepository, None, None]:
    database_url = os.getenv("DATABASE_URL")
    if not database_url:
        raise RuntimeError("DATABASE_URL não configurada")
    yield PostgresImportRepository(database_url)


def storage_root() -> Path:
    configured = os.getenv("IMPORT_STORAGE_DIR")
    root = (
        Path(configured)
        if configured
        else Path(__file__).resolve().parents[2] / "data" / "imports"
    )
    root.mkdir(parents=True, exist_ok=True)
    return root


def configured_organizations() -> set[str] | None:
    configured = os.getenv("VALID_ORGANIZATION_CODES")
    if configured is None:
        return None
    values = {
        value.strip().upper() for value in configured.split(",") if value.strip()
    }
    return values or None


def _validate(path: Path, extension: str) -> None:
    try:
        if extension == ".json":
            with path.open(encoding="utf-8") as stream:
                json.load(stream)
        elif extension == ".csv":
            with path.open(encoding="utf-8-sig", newline="") as stream:
                rows = csv.reader(stream)
                header = next(rows, None)
                if not header or not any(cell.strip() for cell in header):
                    raise ValueError("CSV sem cabeçalho")
        else:
            with path.open("rb") as stream:
                workbook = load_workbook(stream, read_only=True, data_only=True)
                if not workbook.sheetnames:
                    raise ValueError("XLSX sem planilhas")
                workbook.close()
    except (UnicodeError, csv.Error, json.JSONDecodeError, OSError, ValueError) as exc:
        raise ValueError(f"arquivo {extension[1:].upper()} inválido") from exc
    except Exception as exc:
        if extension == ".xlsx":
            raise ValueError("arquivo XLSX inválido") from exc
        raise


def _error(
    repository: ImportRepository,
    execution_id: str,
    http_status: int,
    code: str,
    message: str,
) -> None:
    repository.finish(execution_id, "failed", code)
    logger.warning("import_failed execution_id=%s reason=%s", execution_id, code)
    raise HTTPException(
        status_code=http_status,
        detail={"code": code, "message": message, "execution_id": execution_id},
    )


@router.post(
    "",
    response_model=ImportStatus,
    status_code=status.HTTP_201_CREATED,
    summary="Enviar e registrar um arquivo de entrada",
    responses={
        409: {"description": "Arquivo duplicado por SHA-256"},
        415: {"description": "Extensão não suportada"},
        422: {"description": "Arquivo vazio, inválido ou requisição inválida"},
        500: {"description": "Falha ao preservar o arquivo no storage"},
    },
)
async def upload_import(
    source: Annotated[ImportSource, Form(description="Sistema de origem do arquivo")],
    file: Annotated[UploadFile, File(description="Arquivo XLSX, CSV ou JSON")],
    imported_by: Annotated[str | None, Form(description="Usuário responsável")] = None,
    technical_origin: Annotated[
        str | None, Form(description="Identificador da origem técnica")
    ] = None,
    repository: ImportRepository = Depends(get_repository),
) -> ImportStatus:
    execution_id = str(uuid4())
    actor_type, actor = (
        ("user", imported_by.strip())
        if imported_by and imported_by.strip()
        else ("technical", technical_origin.strip())
        if technical_origin and technical_origin.strip()
        else ("", "")
    )
    if not actor:
        raise HTTPException(
            status_code=422,
            detail="Informe imported_by ou technical_origin",
        )
    repository.start(execution_id, source.value, actor_type, actor)
    logger.info("import_started execution_id=%s source=%s", execution_id, source.value)

    safe_name = Path(file.filename or "").name
    extension = Path(safe_name).suffix.lower()
    if extension not in {".xlsx", ".csv", ".json"}:
        _error(
            repository,
            execution_id,
            415,
            "unsupported_extension",
            "Extensão não suportada",
        )

    temp_path: Path | None = None
    try:
        digest = hashlib.sha256()
        size_bytes = 0
        with NamedTemporaryFile(
            delete=False, dir=storage_root(), suffix=".upload"
        ) as temp:
            temp_path = Path(temp.name)
            while chunk := await file.read(1024 * 1024):
                size_bytes += len(chunk)
                digest.update(chunk)
                temp.write(chunk)
        if size_bytes == 0:
            _error(repository, execution_id, 422, "empty_file", "Arquivo vazio")
        content_hash = digest.hexdigest()
        _validate(temp_path, extension)
        destination = (
            storage_root() / source.name / execution_id / f"original{extension}"
        )
        storage_key = str(destination.relative_to(storage_root()))
        duplicate_of = repository.claim_file(
            execution_id,
            file_name=safe_name,
            extension=extension[1:],
            size_bytes=size_bytes,
            digest=content_hash,
            media_type=file.content_type,
            storage_key=storage_key,
        )
        if duplicate_of:
            repository.finish(execution_id, "duplicate", "duplicate_file", duplicate_of)
            logger.info(
                "import_duplicate execution_id=%s duplicate_of=%s",
                execution_id,
                duplicate_of,
            )
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "duplicate_file",
                    "message": "Arquivo já importado",
                    "execution_id": execution_id,
                    "duplicate_of_execution_id": duplicate_of,
                },
            )
        try:
            destination.parent.mkdir(parents=True, exist_ok=False)
            shutil.move(temp_path, destination)
            temp_path = None
        except OSError as exc:
            destination.unlink(missing_ok=True)
            try:
                destination.parent.rmdir()
            except OSError:
                pass
            repository.abort_claim(execution_id, "storage_error")
            logger.exception("import_storage_failed execution_id=%s", execution_id)
            raise HTTPException(
                status_code=500,
                detail={
                    "code": "storage_error",
                    "message": "Não foi possível preservar o arquivo",
                    "execution_id": execution_id,
                },
            ) from exc
        try:
            report = validate_file(
                destination,
                extension,
                source.value,
                configured_organizations(),
            )
        except Exception:
            logger.exception(
                "import_validation_read_failed execution_id=%s", execution_id
            )
            report = failed_validation_report(
                source.value,
                "validation_read_error",
                "Falha inesperada durante a leitura do arquivo",
            )
        report.update(
            execution_id=execution_id,
            file_name=safe_name,
        )
        for issue in report["issues"]:
            issue["file_name"] = safe_name
        report_path = destination.parent / "validation-report.json"
        report_path.write_text(
            json.dumps(report, ensure_ascii=False, indent=2, default=str) + "\n",
            encoding="utf-8",
        )
        if report["blocking"]:
            repository.mark_validation_failed(execution_id)
            logger.warning(
                "import_validation_blocked execution_id=%s errors=%d warnings=%d",
                execution_id,
                report["error_count"],
                report["warning_count"],
            )
        else:
            repository.mark_completed(execution_id)
        logger.info(
            "import_processed execution_id=%s size_bytes=%d", execution_id, size_bytes
        )
    except HTTPException:
        raise
    except ValueError as exc:
        _error(repository, execution_id, 422, "invalid_file", str(exc))
    finally:
        await file.close()
        if temp_path is not None:
            temp_path.unlink(missing_ok=True)

    result = repository.get(execution_id)
    if result is None:
        raise RuntimeError("Execução recém-criada não encontrada")
    return ImportStatus.model_validate(result)


@router.get(
    "/{execution_id}",
    response_model=ImportStatus,
    summary="Consultar o estado de uma importação",
)
def get_import(
    execution_id: str,
    repository: ImportRepository = Depends(get_repository),
) -> ImportStatus:
    result = repository.get(execution_id)
    if result is None:
        raise HTTPException(
            status_code=404, detail="Execução de importação não encontrada"
        )
    return ImportStatus.model_validate(result)


@router.get(
    "/{execution_id}/validation-report",
    response_model=ValidationReport,
    summary="Visualizar o relatório de validação da execução",
)
def get_validation_report(
    execution_id: str,
    repository: ImportRepository = Depends(get_repository),
) -> ValidationReport:
    execution = repository.get(execution_id)
    if execution is None:
        raise HTTPException(
            status_code=404, detail="Execução de importação não encontrada"
        )
    source = ImportSource(execution["source"])
    report_path = storage_root() / source.name / execution_id / "validation-report.json"
    if not report_path.is_file():
        raise HTTPException(
            status_code=404, detail="Relatório de validação não disponível"
        )
    with report_path.open(encoding="utf-8") as stream:
        return ValidationReport.model_validate(json.load(stream))
