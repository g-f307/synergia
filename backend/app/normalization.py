from __future__ import annotations

import csv
import json
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

RULES_PATH = Path(__file__).with_name("model") / "normalization_rules.json"
REQUIRED_RULES = {
    "column_aliases": dict,
    "identifier_fields": list,
    "date_fields": list,
    "state_map": dict,
    "oqc_flag_map": dict,
}


def _unique_object(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in pairs:
        if key in result:
            raise ValueError(f"Chave duplicada nas regras de normalização: {key}")
        result[key] = value
    return result


def load_normalization_rules(path: Path = RULES_PATH) -> dict[str, Any]:
    try:
        with path.open(encoding="utf-8") as stream:
            rules = json.load(stream, object_pairs_hook=_unique_object)
    except (OSError, json.JSONDecodeError) as exc:
        raise RuntimeError(f"Não foi possível carregar as regras: {path.name}") from exc
    if not isinstance(rules, dict):
        raise ValueError("As regras de normalização devem ser um objeto JSON")
    for name, expected_type in REQUIRED_RULES.items():
        value = rules.get(name)
        if not isinstance(value, expected_type):
            raise ValueError(f"Regra ausente ou inválida: {name}")
    for name in ("column_aliases", "state_map"):
        if not all(
            isinstance(key, str) and isinstance(value, str)
            for key, value in rules[name].items()
        ):
            raise ValueError(f"A regra {name} deve mapear texto para texto")
    if not all(
        isinstance(key, str) and isinstance(value, bool)
        for key, value in rules["oqc_flag_map"].items()
    ):
        raise ValueError("A regra oqc_flag_map deve mapear texto para booleano")
    for name in ("identifier_fields", "date_fields"):
        if not all(isinstance(value, str) for value in rules[name]):
            raise ValueError(f"A regra {name} deve conter apenas textos")
    return rules


NORMALIZATION_RULES = load_normalization_rules()
COLUMN_ALIASES: dict[str, str] = NORMALIZATION_RULES["column_aliases"]
IDENTIFIER_FIELDS = frozenset(NORMALIZATION_RULES["identifier_fields"])
DATE_FIELDS = frozenset(NORMALIZATION_RULES["date_fields"])
STATE_MAP: dict[str, str] = NORMALIZATION_RULES["state_map"]
OQC_FLAG_MAP: dict[str, bool] = NORMALIZATION_RULES["oqc_flag_map"]
BOOLEAN_FIELDS = frozenset(
    {"active", "hold_flag", "rework_flag", "ship_block_flag"}
)


def _without_accents(value: str) -> str:
    return "".join(
        character
        for character in unicodedata.normalize("NFKD", value)
        if not unicodedata.combining(character)
    )


def normalize_column_name(value: Any) -> str:
    text = _without_accents(str(value or "").strip().lower())
    normalized = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    return COLUMN_ALIASES.get(normalized, normalized)


def _text(value: Any) -> str:
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float | Decimal):
        try:
            decimal = Decimal(str(value))
        except InvalidOperation:
            return str(value)
        return format(decimal, "f")
    return str(value)


def normalize_identifier(value: Any) -> str | None:
    if value in (None, ""):
        return None
    text = _text(value).strip()
    if re.fullmatch(r"[+-]?\d+(?:\.\d+)?[Ee][+-]?\d+", text):
        try:
            text = format(Decimal(text), "f")
        except InvalidOperation:
            pass
    text = text.upper()
    text = re.sub(r"\s*([./-])\s*", r"\1", text)
    return re.sub(r"\s+", "", text)


def normalize_date(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        if value.time() == datetime.min.time():
            return value.date().isoformat()
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for date_format in (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
    ):
        try:
            parsed = datetime.strptime(text, date_format)
            return (
                parsed.date().isoformat()
                if parsed.time() == datetime.min.time()
                else parsed.isoformat()
            )
        except ValueError:
            continue
    return text


def _lookup_key(value: Any) -> str:
    return re.sub(r"\s+", " ", _without_accents(str(value).strip().lower()))


def normalize_state(value: Any) -> tuple[str | None, bool]:
    if value in (None, ""):
        return None, True
    key = _lookup_key(value)
    return STATE_MAP.get(key, key.replace(" ", "_")), key in STATE_MAP


def normalize_oqc_flag(value: Any) -> tuple[bool | None, bool]:
    if value in (None, ""):
        return None, True
    if isinstance(value, bool):
        return value, True
    key = _lookup_key(value)
    return OQC_FLAG_MAP.get(key), key in OQC_FLAG_MAP


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, str | int | float | bool):
        return value
    if isinstance(value, date | datetime):
        return value.isoformat()
    return str(value)


def _read_tables(path: Path, extension: str):
    if extension == ".csv":
        with path.open(encoding="utf-8-sig", newline="") as stream:
            rows = list(csv.reader(stream, strict=True))
        if rows:
            yield "CSV", rows[0], rows[1:]
        return
    if extension == ".json":
        with path.open(encoding="utf-8") as stream:
            document = json.load(stream)
        records = document if isinstance(document, list) else [document]
        headers = list(dict.fromkeys(key for record in records for key in record))
        yield (
            "JSON",
            headers,
            [[record.get(key) for key in headers] for record in records],
        )
        return
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        for worksheet in workbook.worksheets:
            rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            if rows and any(value not in (None, "") for value in rows[0]):
                yield worksheet.title, rows[0], rows[1:]
    finally:
        workbook.close()


def _normalize_value(field: str, value: Any) -> tuple[Any, str, str | None]:
    if field in IDENTIFIER_FIELDS:
        return normalize_identifier(value), "identifier_to_text", None
    if field in DATE_FIELDS or field.endswith("_date") or field.endswith("_at"):
        return normalize_date(value), "date_to_iso8601", None
    if field == "status":
        normalized, known = normalize_state(value)
        return normalized, "state_mapping", None if known else "unknown_state"
    if field == "oqc_flag":
        normalized, known = normalize_oqc_flag(value)
        return normalized, "oqc_flag_mapping", None if known else "unknown_oqc_flag"
    if field in BOOLEAN_FIELDS:
        normalized, known = normalize_oqc_flag(value)
        return (
            normalized,
            "boolean_flag_mapping",
            None if known else "unknown_boolean_flag",
        )
    if isinstance(value, str):
        return value.strip(), "trim_whitespace", None
    return _json_value(value), "preserve_value", None


def normalize_tables(
    tables: list[tuple[str, list[Any], list[list[Any]]]],
    source: str,
    eligible_rows: set[tuple[str, int]] | None = None,
) -> dict[str, Any]:
    records: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    for sheet, raw_headers, rows in tables:
        headers = [normalize_column_name(header) for header in raw_headers]
        for row_number, row in enumerate(rows, 2):
            row_number = getattr(row, "row_number", row_number)
            if not any(value not in (None, "") for value in row):
                continue
            if eligible_rows is not None and (sheet, row_number) not in eligible_rows:
                continue
            values: dict[str, Any] = {}
            originals: dict[str, Any] = {}
            transformations: list[dict[str, Any]] = []
            for index, field in enumerate(headers):
                original = row[index] if index < len(row) else None
                normalized, operation, issue_code = _normalize_value(field, original)
                source_column = str(raw_headers[index])
                values[field] = normalized
                originals[field] = _json_value(original)
                transformations.append(
                    {
                        "field": field,
                        "source_column": source_column,
                        "original_value": _json_value(original),
                        "normalized_value": normalized,
                        "operation": operation,
                    }
                )
                if issue_code:
                    issues.append(
                        {
                            "severity": "warning",
                            "code": issue_code,
                            "sheet": sheet,
                            "row": row_number,
                            "column": source_column,
                            "field": field,
                            "original_value": _json_value(original),
                            "reason": "Valor sem mapeamento conhecido",
                        }
                    )
            records.append(
                {
                    "source": source,
                    "sheet": sheet,
                    "row": row_number,
                    "values": values,
                    "original_values": originals,
                    "transformations": transformations,
                }
            )
    return {
        "source": source,
        "record_count": len(records),
        "warning_count": len(issues),
        "issues": issues,
        "records": records,
    }


def normalize_file(path: Path, extension: str, source: str) -> dict[str, Any]:
    return normalize_tables(list(_read_tables(path, extension)), source)
